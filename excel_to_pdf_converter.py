
#!/usr/bin/env python3
"""
Enhanced Excel to PDF Converter
Converts Excel files to PDF while preserving formatting, layout, images, tables, and charts.
Uses multiple methods for optimal conversion quality.
"""

import os
import tempfile
import traceback
import shutil
from pathlib import Path
from typing import Dict, Any, Optional
import uuid
import subprocess

try:
    import openpyxl
    from openpyxl import load_workbook
    print("openpyxl successfully imported")
except ImportError as e:
    print(f"Warning: openpyxl not available - {e}")
    openpyxl = None

try:
    import xlrd
    print("xlrd successfully imported")
except ImportError as e:
    print(f"Warning: xlrd not available - {e}")
    xlrd = None

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    print("ReportLab successfully imported")
except ImportError as e:
    print(f"Warning: ReportLab not available - {e}")
    SimpleDocTemplate = None

try:
    import pandas as pd
    print("pandas successfully imported")
except ImportError as e:
    print(f"Warning: pandas not available - {e}")
    pd = None

try:
    from PIL import Image
    print("Pillow successfully imported")
except ImportError as e:
    print(f"Warning: Pillow not available - {e}")
    Image = None

class ExcelToPDFConverter:
    """Advanced Excel to PDF converter with comprehensive formatting preservation"""
    
    def __init__(self, output_dir: str = None):
        self.output_dir = output_dir or tempfile.mkdtemp()
        Path(self.output_dir).mkdir(parents=True, exist_ok=True)
        
    def convert_excel_to_pdf(self, excel_path: str, output_filename: str = None) -> Dict[str, Any]:
        """
        Convert Excel to PDF with enhanced formatting preservation and multiple fallback methods
        """
        try:
            print(f"Starting enhanced Excel to PDF conversion: {excel_path}")
            
            # Validate input file
            if not os.path.exists(excel_path):
                return {"success": False, "error": "Excel file not found"}
                
            if not (excel_path.lower().endswith('.xlsx') or excel_path.lower().endswith('.xls')):
                return {"success": False, "error": "File must be an Excel file (.xlsx or .xls)"}
            
            # Generate output filename
            if not output_filename:
                base_name = Path(excel_path).stem
                output_filename = f"{base_name}_converted.pdf"
            elif not output_filename.endswith('.pdf'):
                output_filename += '.pdf'
                
            output_path = os.path.join(self.output_dir, output_filename)
            
            # Try multiple conversion methods in order of effectiveness
            methods = [
                ("LibreOffice conversion", self._convert_with_libreoffice),
                ("ReportLab advanced conversion", self._convert_with_reportlab),
                ("Pandas HTML to PDF", self._convert_with_pandas_html),
                ("Basic text extraction", self._convert_with_basic_extraction)
            ]
            
            for method_name, method_func in methods:
                try:
                    print(f"Attempting {method_name}...")
                    if method_func(excel_path, output_path):
                        print(f"{method_name} succeeded")
                        
                        # Get file metadata
                        metadata = self._get_file_metadata(output_path, excel_path)
                        
                        return {
                            "success": True, 
                            "output_path": output_path,
                            "filename": output_filename,
                            "method": method_name.lower().replace(" ", "_"),
                            "message": f"Conversion completed using {method_name}",
                            "metadata": metadata
                        }
                except Exception as e:
                    print(f"{method_name} failed: {e}")
                    continue
            
            return {"success": False, "error": "All conversion methods failed"}
            
        except Exception as e:
            error_msg = f"Error in convert_excel_to_pdf: {str(e)}"
            print(f"Error in convert_excel_to_pdf: {error_msg}")
            print(traceback.format_exc())
            return {"success": False, "error": error_msg}

    def _convert_with_libreoffice(self, excel_path: str, output_path: str) -> bool:
        """Convert Excel to PDF using LibreOffice (best quality)"""
        try:
            print("Attempting LibreOffice conversion...")
            
            # Create temporary directory for LibreOffice
            temp_output_dir = tempfile.mkdtemp()
            
            # LibreOffice command
            cmd = [
                'libreoffice', 
                '--headless', 
                '--convert-to', 'pdf',
                '--outdir', temp_output_dir,
                excel_path
            ]
            
            print(f"Running command: {' '.join(cmd)}")
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=300)
            
            if result.returncode == 0:
                # Find the generated PDF file
                base_name = Path(excel_path).stem
                temp_pdf_path = os.path.join(temp_output_dir, f"{base_name}.pdf")
                
                if os.path.exists(temp_pdf_path):
                    shutil.move(temp_pdf_path, output_path)
                    shutil.rmtree(temp_output_dir, ignore_errors=True)
                    print("LibreOffice conversion successful")
                    return True
                else:
                    print(f"LibreOffice output file not found at {temp_pdf_path}")
            else:
                print(f"LibreOffice failed with return code {result.returncode}")
                print(f"Error output: {result.stderr}")
            
            shutil.rmtree(temp_output_dir, ignore_errors=True)
            return False
            
        except subprocess.TimeoutExpired:
            print("LibreOffice conversion timed out")
            return False
        except Exception as e:
            print(f"LibreOffice conversion error: {e}")
            return False

    def _convert_with_reportlab(self, excel_path: str, output_path: str) -> bool:
        """Convert Excel to PDF using ReportLab with advanced formatting"""
        try:
            if not openpyxl or not SimpleDocTemplate:
                return False
                
            print("Attempting ReportLab conversion...")
            
            # Load workbook
            wb = load_workbook(excel_path, data_only=True)
            
            # Create PDF document
            doc = SimpleDocTemplate(output_path, pagesize=A4)
            story = []
            styles = getSampleStyleSheet()
            
            # Process each worksheet
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Add sheet title
                title = Paragraph(f"<b>{sheet_name}</b>", styles['Heading1'])
                story.append(title)
                story.append(Spacer(1, 12))
                
                # Get data range
                if ws.max_row > 0 and ws.max_column > 0:
                    data = []
                    
                    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 100), 
                                          min_col=1, max_col=min(ws.max_column, 20)):
                        row_data = []
                        for cell in row:
                            value = cell.value
                            if value is None:
                                value = ""
                            else:
                                value = str(value)
                            row_data.append(value)
                        data.append(row_data)
                    
                    if data:
                        # Create table
                        table = Table(data)
                        
                        # Style the table
                        table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 10),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black),
                            ('FONTSIZE', (0, 1), (-1, -1), 8),
                        ]))
                        
                        story.append(table)
                        story.append(Spacer(1, 20))
            
            # Build PDF
            doc.build(story)
            print("ReportLab conversion successful")
            return True
            
        except Exception as e:
            print(f"ReportLab conversion error: {e}")
            return False

    def _convert_with_pandas_html(self, excel_path: str, output_path: str) -> bool:
        """Convert Excel to PDF via HTML using pandas"""
        try:
            if not pd:
                return False
                
            print("Attempting pandas HTML conversion...")
            
            # Read Excel file
            if excel_path.lower().endswith('.xlsx'):
                sheets = pd.read_excel(excel_path, sheet_name=None, engine='openpyxl')
            else:
                sheets = pd.read_excel(excel_path, sheet_name=None, engine='xlrd')
            
            # Create HTML content
            html_content = """
            <html>
            <head>
                <style>
                    body { font-family: Arial, sans-serif; margin: 20px; }
                    table { border-collapse: collapse; width: 100%; margin-bottom: 30px; }
                    th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
                    th { background-color: #f2f2f2; font-weight: bold; }
                    h2 { color: #333; border-bottom: 2px solid #333; }
                </style>
            </head>
            <body>
            """
            
            for sheet_name, df in sheets.items():
                html_content += f"<h2>{sheet_name}</h2>\n"
                html_content += df.to_html(escape=False, index=False, classes='table')
                html_content += "<br><br>\n"
            
            html_content += "</body></html>"
            
            # Save HTML to temp file
            temp_html = os.path.join(self.output_dir, f"temp_{uuid.uuid4().hex}.html")
            with open(temp_html, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            # Convert HTML to PDF using wkhtmltopdf if available
            try:
                cmd = ['wkhtmltopdf', '--page-size', 'A4', '--orientation', 'Portrait', 
                       '--margin-top', '0.75in', '--margin-right', '0.75in', 
                       '--margin-bottom', '0.75in', '--margin-left', '0.75in',
                       temp_html, output_path]
                
                result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
                
                if result.returncode == 0 and os.path.exists(output_path):
                    os.remove(temp_html)
                    print("Pandas HTML conversion successful")
                    return True
                    
            except (subprocess.TimeoutExpired, FileNotFoundError):
                pass
            
            # Cleanup
            if os.path.exists(temp_html):
                os.remove(temp_html)
            
            return False
            
        except Exception as e:
            print(f"Pandas HTML conversion error: {e}")
            return False

    def _convert_with_basic_extraction(self, excel_path: str, output_path: str) -> bool:
        """Basic fallback conversion using simple text extraction"""
        try:
            if not SimpleDocTemplate or not openpyxl:
                return False
                
            print("Attempting basic text extraction...")
            
            # Load workbook
            wb = load_workbook(excel_path, data_only=True)
            
            # Create simple PDF
            doc = SimpleDocTemplate(output_path, pagesize=letter)
            story = []
            styles = getSampleStyleSheet()
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Add sheet title
                title = Paragraph(f"<b>{sheet_name}</b>", styles['Heading1'])
                story.append(title)
                story.append(Spacer(1, 12))
                
                # Extract text content
                content = []
                for row in ws.iter_rows(values_only=True, max_row=50, max_col=10):
                    row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        content.append(row_text)
                
                # Add content as paragraphs
                for line in content:
                    if line.strip():
                        para = Paragraph(line, styles['Normal'])
                        story.append(para)
                        story.append(Spacer(1, 6))
                
                story.append(Spacer(1, 20))
            
            doc.build(story)
            print("Basic text extraction successful")
            return True
            
        except Exception as e:
            print(f"Basic extraction error: {e}")
            return False

    def _get_file_metadata(self, output_path: str, input_path: str) -> Dict[str, Any]:
        """Get metadata about the converted file"""
        metadata = {}
        
        try:
            if os.path.exists(output_path):
                metadata['file_size'] = os.path.getsize(output_path)
            
            # Try to get Excel metadata
            if openpyxl and input_path.lower().endswith('.xlsx'):
                wb = load_workbook(input_path, data_only=True)
                metadata['sheets'] = len(wb.sheetnames)
                metadata['sheet_names'] = wb.sheetnames
                
                # Count total cells with data
                total_cells = 0
                for ws in wb.worksheets:
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.value is not None:
                                total_cells += 1
                metadata['total_cells'] = total_cells
                
        except Exception as e:
            print(f"Error getting metadata: {e}")
            
        return metadata

def convert_excel_file(excel_path: str, output_dir: str = None) -> Dict[str, Any]:
    """
    Main function to convert Excel to PDF with enhanced capabilities
    """
    converter = ExcelToPDFConverter(output_dir)
    return converter.convert_excel_to_pdf(excel_path)

def test_converter():
    """Test function for the Excel to PDF converter"""
    converter = ExcelToPDFConverter()
    
    test_excel = "test_sample.xlsx"
    if not os.path.exists(test_excel):
        print("No test Excel found, skipping test")
        return
    
    result = converter.convert_excel_to_pdf(test_excel)
    print(f"Converter test result: {result}")

if __name__ == "__main__":
    test_converter()
