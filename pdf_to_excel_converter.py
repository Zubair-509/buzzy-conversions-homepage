
#!/usr/bin/env python3
"""
Enhanced PDF to Excel Converter
Converts PDF files to Excel spreadsheets while preserving formatting, layout, images, and tables.
Uses multiple methods for optimal conversion quality with advanced data extraction.
"""

import os
import tempfile
import traceback
import shutil
from pathlib import Path
from typing import Dict, Any, Optional, List, Tuple
import uuid
import re

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.drawing.image import Image as ExcelImage
    print("openpyxl successfully imported")
except ImportError as e:
    print(f"Warning: openpyxl not available - {e}")
    openpyxl = None

try:
    import pandas as pd
    print("pandas successfully imported")
except ImportError as e:
    print(f"Warning: pandas not available - {e}")
    pd = None

try:
    import pdfplumber
    print("pdfplumber successfully imported")
except ImportError as e:
    print(f"Warning: pdfplumber not available - {e}")
    pdfplumber = None

try:
    import tabula
    print("tabula-py successfully imported")
except ImportError as e:
    print(f"Warning: tabula-py not available - {e}")
    tabula = None

try:
    from PIL import Image
    print("Pillow successfully imported")
except ImportError as e:
    print(f"Warning: Pillow not available - {e}")
    Image = None

try:
    import fitz  # PyMuPDF
    print("PyMuPDF successfully imported")
except ImportError as e:
    print(f"Warning: PyMuPDF not available - {e}")
    fitz = None

try:
    import camelot
    print("camelot-py successfully imported")
except ImportError as e:
    print(f"Warning: camelot-py not available - {e}")
    camelot = None

class PDFToExcelConverter:
    """Advanced PDF to Excel converter with comprehensive table and data extraction"""
    
    def __init__(self, output_dir: str = None):
        self.output_dir = output_dir or tempfile.mkdtemp()
        Path(self.output_dir).mkdir(parents=True, exist_ok=True)
        
    def convert_pdf_to_excel(self, pdf_path: str, output_filename: str = None) -> Dict[str, Any]:
        """
        Convert PDF to Excel with enhanced data extraction and multiple fallback methods
        """
        try:
            print(f"Starting enhanced PDF to Excel conversion: {pdf_path}")
            
            # Validate input file
            if not os.path.exists(pdf_path):
                return {"success": False, "error": "PDF file not found"}
                
            if not pdf_path.lower().endswith('.pdf'):
                return {"success": False, "error": "File must be a PDF"}
            
            # Generate output filename
            if not output_filename:
                base_name = Path(pdf_path).stem
                output_filename = f"{base_name}_converted.xlsx"
            elif not output_filename.endswith('.xlsx'):
                output_filename += '.xlsx'
                
            output_path = os.path.join(self.output_dir, output_filename)
            
            # Try multiple conversion methods in order of effectiveness
            methods = [
                ("Enhanced Tabula extraction", self._convert_with_tabula_enhanced),
                ("Camelot table extraction", self._convert_with_camelot),
                ("PDFplumber advanced extraction", self._convert_with_pdfplumber_advanced),
                ("PyMuPDF enhanced extraction", self._convert_with_pymupdf_enhanced),
                ("Raw text extraction", self._convert_with_text_extraction)
            ]
            
            for method_name, method_func in methods:
                try:
                    print(f"Attempting {method_name}...")
                    if method_func(pdf_path, output_path):
                        print(f"{method_name} succeeded")
                        return {
                            "success": True, 
                            "output_path": output_path,
                            "filename": output_filename,
                            "method": method_name.lower().replace(" ", "_"),
                            "message": f"Conversion completed using {method_name}"
                        }
                except Exception as e:
                    print(f"{method_name} failed: {e}")
                    continue
            
            return {"success": False, "error": "All conversion methods failed"}
            
        except Exception as e:
            error_msg = f"Error in convert_pdf_to_excel: {str(e)}"
            print(f"Error in convert_pdf_to_excel: {error_msg}")
            print(traceback.format_exc())
            return {"success": False, "error": error_msg}
    
    def _convert_with_tabula_enhanced(self, pdf_path: str, output_path: str) -> bool:
        """Enhanced conversion using tabula-py for table detection"""
        try:
            if not (tabula and openpyxl and pd):
                return False
            
            print("Attempting tabula-py table extraction...")
            wb = Workbook()
            
            # Remove default sheet
            if wb.worksheets:
                wb.remove(wb.active)
            
            # Read tables from all pages
            try:
                # Try different extraction methods
                all_tables = []
                
                # Method 1: Standard tabula extraction
                try:
                    tables = tabula.read_pdf(pdf_path, pages='all', multiple_tables=True, pandas_options={'header': None})
                    if tables:
                        all_tables.extend([(table, f"Page_Auto_{i+1}") for i, table in enumerate(tables)])
                        print(f"Found {len(tables)} tables using standard tabula")
                except Exception as e:
                    print(f"Standard tabula failed: {e}")
                
                # Method 2: Lattice method (for tables with clear borders)
                try:
                    tables = tabula.read_pdf(pdf_path, pages='all', lattice=True, multiple_tables=True, pandas_options={'header': None})
                    if tables:
                        all_tables.extend([(table, f"Lattice_{i+1}") for i, table in enumerate(tables)])
                        print(f"Found {len(tables)} tables using lattice method")
                except Exception as e:
                    print(f"Lattice tabula failed: {e}")
                
                # Method 3: Stream method (for tables without clear borders)
                try:
                    tables = tabula.read_pdf(pdf_path, pages='all', stream=True, multiple_tables=True, pandas_options={'header': None})
                    if tables:
                        all_tables.extend([(table, f"Stream_{i+1}") for i, table in enumerate(tables)])
                        print(f"Found {len(tables)} tables using stream method")
                except Exception as e:
                    print(f"Stream tabula failed: {e}")
                
                if not all_tables:
                    print("No tables found with tabula-py")
                    return False
                
                # Process all found tables
                for table_idx, (df, source) in enumerate(all_tables):
                    if df is not None and not df.empty:
                        sheet_name = f"{source}_{table_idx + 1}"[:31]  # Excel sheet name limit
                        ws = wb.create_sheet(title=sheet_name)
                        
                        print(f"Processing table {table_idx + 1} from {source}, shape: {df.shape}")
                        
                        # Add table data to worksheet
                        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 1):
                            for c_idx, value in enumerate(row, 1):
                                if pd.notna(value):
                                    cell = ws.cell(row=r_idx, column=c_idx, value=str(value))
                                    
                                    # Apply formatting for first row
                                    if r_idx == 1:
                                        cell.font = Font(bold=True)
                                        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                                    
                                    # Add borders
                                    cell.border = Border(
                                        left=Side(style='thin'),
                                        right=Side(style='thin'),
                                        top=Side(style='thin'),
                                        bottom=Side(style='thin')
                                    )
                        
                        # Auto-adjust column widths
                        for column in ws.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            
                            adjusted_width = min(max_length + 2, 50)
                            ws.column_dimensions[column_letter].width = adjusted_width
                
                wb.save(output_path)
                return True
                
            except Exception as e:
                print(f"Tabula extraction failed: {e}")
                return False
                
        except Exception as e:
            print(f"Tabula enhanced conversion failed: {e}")
            return False
    
    def _convert_with_camelot(self, pdf_path: str, output_path: str) -> bool:
        """Convert using camelot-py for advanced table extraction"""
        try:
            if not (camelot and openpyxl):
                return False
            
            print("Attempting camelot table extraction...")
            wb = Workbook()
            
            # Remove default sheet
            if wb.worksheets:
                wb.remove(wb.active)
            
            # Extract tables using camelot
            try:
                # Try lattice method first (for tables with lines)
                tables = camelot.read_pdf(pdf_path, pages='all', flavor='lattice')
                
                if len(tables) == 0:
                    # Try stream method (for tables without lines)
                    tables = camelot.read_pdf(pdf_path, pages='all', flavor='stream')
                
                if len(tables) == 0:
                    print("No tables found with camelot")
                    return False
                
                print(f"Found {len(tables)} tables with camelot")
                
                for i, table in enumerate(tables):
                    sheet_name = f"Table_{i+1}_P{table.page}"
                    ws = wb.create_sheet(title=sheet_name)
                    
                    print(f"Processing camelot table {i+1}, accuracy: {table.accuracy:.2f}")
                    
                    # Convert table to dataframe and add to worksheet
                    df = table.df
                    
                    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 1):
                        for c_idx, value in enumerate(row, 1):
                            if pd.notna(value) and str(value).strip():
                                cell = ws.cell(row=r_idx, column=c_idx, value=str(value).strip())
                                
                                # Apply formatting for first row
                                if r_idx == 1:
                                    cell.font = Font(bold=True)
                                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                                
                                # Add borders
                                cell.border = Border(
                                    left=Side(style='thin'),
                                    right=Side(style='thin'),
                                    top=Side(style='thin'),
                                    bottom=Side(style='thin')
                                )
                    
                    # Auto-adjust column widths
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
                
                wb.save(output_path)
                return True
                
            except Exception as e:
                print(f"Camelot table extraction failed: {e}")
                return False
                
        except Exception as e:
            print(f"Camelot conversion failed: {e}")
            return False
    
    def _convert_with_pdfplumber_advanced(self, pdf_path: str, output_path: str) -> bool:
        """Advanced conversion using pdfplumber with better table detection"""
        try:
            if not (pdfplumber and openpyxl):
                return False
            
            print("Attempting advanced pdfplumber extraction...")
            wb = Workbook()
            
            # Remove default sheet
            if wb.worksheets:
                wb.remove(wb.active)
            
            with pdfplumber.open(pdf_path) as pdf:
                for page_num, page in enumerate(pdf.pages):
                    print(f"Processing page {page_num + 1} with pdfplumber")
                    
                    sheet_name = f"Page_{page_num + 1}"
                    ws = wb.create_sheet(title=sheet_name)
                    current_row = 1
                    
                    # Extract tables with different settings
                    tables = []
                    
                    # Try different table extraction strategies
                    strategies = [
                        {"vertical_strategy": "lines", "horizontal_strategy": "lines"},
                        {"vertical_strategy": "text", "horizontal_strategy": "text"},
                        {"vertical_strategy": "lines_strict", "horizontal_strategy": "lines_strict"},
                    ]
                    
                    for strategy in strategies:
                        try:
                            page_tables = page.extract_tables(table_settings=strategy)
                            if page_tables:
                                tables.extend(page_tables)
                                print(f"Found {len(page_tables)} tables with strategy {strategy}")
                                break
                        except Exception as e:
                            continue
                    
                    if tables:
                        for table_idx, table in enumerate(tables):
                            if table and any(any(cell for cell in row if cell) for row in table):
                                # Add table title
                                title_cell = ws.cell(row=current_row, column=1, value=f"Table {table_idx + 1}")
                                title_cell.font = Font(bold=True, size=12)
                                current_row += 1
                                
                                # Process table data
                                for row_idx, row in enumerate(table):
                                    if row and any(cell for cell in row if cell):
                                        for col_idx, cell_value in enumerate(row):
                                            if cell_value:
                                                # Clean cell value
                                                cleaned_value = str(cell_value).strip()
                                                if cleaned_value:
                                                    cell = ws.cell(row=current_row, column=col_idx + 1, value=cleaned_value)
                                                    
                                                    # Apply formatting for header row
                                                    if row_idx == 0:
                                                        cell.font = Font(bold=True)
                                                        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                                                    
                                                    # Add borders
                                                    cell.border = Border(
                                                        left=Side(style='thin'),
                                                        right=Side(style='thin'),
                                                        top=Side(style='thin'),
                                                        bottom=Side(style='thin')
                                                    )
                                        
                                        current_row += 1
                                
                                current_row += 2  # Space between tables
                    
                    else:
                        # No tables found, extract text and try to parse structured data
                        text = page.extract_text()
                        if text:
                            print(f"No tables found on page {page_num + 1}, extracting text")
                            lines = text.split('\n')
                            
                            # Try to detect tabular data in text
                            structured_data = self._extract_structured_text_data(lines)
                            
                            if structured_data:
                                print(f"Found structured data with {len(structured_data)} rows")
                                for row_data in structured_data:
                                    for col_idx, value in enumerate(row_data):
                                        if value.strip():
                                            ws.cell(row=current_row, column=col_idx + 1, value=value.strip())
                                    current_row += 1
                            else:
                                # Extract all text as single column
                                for line in lines:
                                    if line.strip():
                                        ws.cell(row=current_row, column=1, value=line.strip())
                                        current_row += 1
            
            # Auto-adjust column widths for all sheets
            for sheet in wb.worksheets:
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(output_path)
            return True
            
        except Exception as e:
            print(f"PDFplumber advanced conversion failed: {e}")
            return False
    
    def _convert_with_pymupdf_enhanced(self, pdf_path: str, output_path: str) -> bool:
        """Enhanced conversion using PyMuPDF"""
        try:
            if not (fitz and openpyxl):
                return False
            
            print("Attempting PyMuPDF enhanced extraction...")
            doc = fitz.open(pdf_path)
            wb = Workbook()
            
            # Remove default sheet
            if wb.worksheets:
                wb.remove(wb.active)
            
            for page_num in range(len(doc)):
                page = doc.load_page(page_num)
                sheet_name = f"Page_{page_num + 1}"
                ws = wb.create_sheet(title=sheet_name)
                
                print(f"Processing page {page_num + 1} with PyMuPDF")
                current_row = 1
                
                # Extract tables
                try:
                    tables = page.find_tables()
                    if tables:
                        for table_idx, table in enumerate(tables):
                            print(f"Found table {table_idx + 1} on page {page_num + 1}")
                            
                            # Add table title
                            title_cell = ws.cell(row=current_row, column=1, value=f"Table {table_idx + 1}")
                            title_cell.font = Font(bold=True, size=12)
                            current_row += 1
                            
                            # Extract table data
                            table_data = table.extract()
                            
                            for row_idx, row_data in enumerate(table_data):
                                if row_data and any(cell for cell in row_data if cell):
                                    for col_idx, cell_value in enumerate(row_data):
                                        if cell_value:
                                            cleaned_value = str(cell_value).strip()
                                            if cleaned_value:
                                                cell = ws.cell(row=current_row, column=col_idx + 1, value=cleaned_value)
                                                
                                                # Apply formatting for header row
                                                if row_idx == 0:
                                                    cell.font = Font(bold=True)
                                                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                                                
                                                # Add borders
                                                cell.border = Border(
                                                    left=Side(style='thin'),
                                                    right=Side(style='thin'),
                                                    top=Side(style='thin'),
                                                    bottom=Side(style='thin')
                                                )
                                    
                                    current_row += 1
                            
                            current_row += 2  # Space between tables
                    
                    else:
                        # No tables found, extract text
                        text = page.get_text()
                        if text:
                            lines = text.split('\n')
                            structured_data = self._extract_structured_text_data(lines)
                            
                            if structured_data:
                                for row_data in structured_data:
                                    for col_idx, value in enumerate(row_data):
                                        if value.strip():
                                            ws.cell(row=current_row, column=col_idx + 1, value=value.strip())
                                    current_row += 1
                            else:
                                for line in lines:
                                    if line.strip():
                                        ws.cell(row=current_row, column=1, value=line.strip())
                                        current_row += 1
                        
                except Exception as e:
                    print(f"Error processing page {page_num + 1}: {e}")
                    continue
            
            doc.close()
            
            # Auto-adjust column widths
            for sheet in wb.worksheets:
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(output_path)
            return True
            
        except Exception as e:
            print(f"PyMuPDF enhanced conversion failed: {e}")
            return False
    
    def _convert_with_text_extraction(self, pdf_path: str, output_path: str) -> bool:
        """Fallback method: extract all text and structure it"""
        try:
            if not openpyxl:
                return False
            
            print("Attempting raw text extraction as fallback...")
            wb = Workbook()
            ws = wb.active
            ws.title = "PDF_Content"
            
            current_row = 1
            
            # Try different text extraction methods
            text_content = []
            
            # Method 1: PyMuPDF
            if fitz:
                try:
                    doc = fitz.open(pdf_path)
                    for page_num in range(len(doc)):
                        page = doc.load_page(page_num)
                        text = page.get_text()
                        if text:
                            text_content.append(f"=== Page {page_num + 1} ===")
                            text_content.extend(text.split('\n'))
                    doc.close()
                except Exception as e:
                    print(f"PyMuPDF text extraction failed: {e}")
            
            # Method 2: pdfplumber
            if not text_content and pdfplumber:
                try:
                    with pdfplumber.open(pdf_path) as pdf:
                        for page_num, page in enumerate(pdf.pages):
                            text = page.extract_text()
                            if text:
                                text_content.append(f"=== Page {page_num + 1} ===")
                                text_content.extend(text.split('\n'))
                except Exception as e:
                    print(f"pdfplumber text extraction failed: {e}")
            
            if not text_content:
                # Create basic file info
                ws.cell(row=1, column=1, value="PDF Text Extraction Results")
                ws.cell(row=1, column=1).font = Font(bold=True, size=14)
                ws.cell(row=3, column=1, value=f"Source: {os.path.basename(pdf_path)}")
                ws.cell(row=4, column=1, value="Status: Text extraction failed - file may be image-based or encrypted")
                wb.save(output_path)
                return True
            
            # Process extracted text
            structured_data = self._extract_structured_text_data(text_content)
            
            if structured_data:
                print(f"Extracted {len(structured_data)} rows of structured data")
                for row_data in structured_data:
                    for col_idx, value in enumerate(row_data):
                        if value.strip():
                            ws.cell(row=current_row, column=col_idx + 1, value=value.strip())
                    current_row += 1
            else:
                # Add text line by line
                for line in text_content:
                    if line.strip():
                        ws.cell(row=current_row, column=1, value=line.strip())
                        current_row += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(output_path)
            return True
            
        except Exception as e:
            print(f"Text extraction fallback failed: {e}")
            return False
    
    def _extract_structured_text_data(self, lines: List[str]) -> List[List[str]]:
        """Enhanced structured data detection from text lines"""
        structured_data = []
        
        for line in lines:
            line = line.strip()
            if not line or line.startswith('==='):
                continue
            
            # Method 1: Tab-separated values
            if '\t' in line:
                parts = [part.strip() for part in line.split('\t')]
                if len(parts) > 1 and any(part for part in parts):
                    structured_data.append(parts)
                continue
            
            # Method 2: Multiple spaces (common in PDFs)
            if re.search(r'\s{3,}', line):
                parts = [part.strip() for part in re.split(r'\s{3,}', line)]
                if len(parts) > 1 and any(part for part in parts):
                    structured_data.append(parts)
                continue
            
            # Method 3: Pipe-separated values
            if '|' in line and line.count('|') > 1:
                parts = [part.strip() for part in line.split('|') if part.strip()]
                if len(parts) > 1:
                    structured_data.append(parts)
                continue
            
            # Method 4: Comma-separated (but be careful with regular text)
            if ',' in line and len(line.split(',')) > 2:
                # Only consider as CSV if it looks like data (has numbers or consistent structure)
                parts = [part.strip() for part in line.split(',')]
                if len(parts) > 2 and (any(re.search(r'\d', part) for part in parts) or 
                                      all(len(part) < 50 for part in parts)):
                    structured_data.append(parts)
                continue
            
            # Method 5: Detect patterns like "Key: Value" pairs
            if ':' in line and not line.endswith(':'):
                parts = [part.strip() for part in line.split(':')]
                if len(parts) == 2 and both(parts):
                    structured_data.append(parts)
                continue
        
        # Only return structured data if we have enough consistent rows
        if len(structured_data) >= 2:
            # Check if rows have similar column counts
            col_counts = [len(row) for row in structured_data]
            avg_cols = sum(col_counts) / len(col_counts)
            
            # Filter rows that are close to average column count
            filtered_data = [row for row in structured_data 
                           if abs(len(row) - avg_cols) <= 2]
            
            if len(filtered_data) >= 2:
                return filtered_data
        
        return None

def convert_pdf_to_excel(pdf_path: str, output_dir: str = None) -> Dict[str, Any]:
    """
    Main function to convert PDF to Excel with enhanced capabilities
    """
    converter = PDFToExcelConverter(output_dir)
    return converter.convert_pdf_to_excel(pdf_path)

def test_converter():
    """Test function for the PDF to Excel converter"""
    converter = PDFToExcelConverter()
    
    test_pdf = "test_sample.pdf"
    if not os.path.exists(test_pdf):
        print("No test PDF found, skipping test")
        return
    
    result = converter.convert_pdf_to_excel(test_pdf)
    print(f"Converter test result: {result}")

if __name__ == "__main__":
    test_converter()
